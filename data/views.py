import json
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import *
from .models import *
from uuid import uuid4
def query_dict_to_json(query):
    json_data = {}
    for dat in query:
        json_data[dat] = json.loads(query[dat])
    return json_data

class GetNote(generics.RetrieveAPIView):
    serializer_class = NoteSerializer
    def get_object(self,*args,**kwargs):
        note = Note.objects.filter(uid=self.kwargs['uid'], is_viewed=False)
        if note.exists():
            note_obj = note[0]
            if not note_obj.is_forever:
                note_obj.is_viewed = True
                note_obj.save()
            return note_obj
        else:
            return None

class Save(APIView):
    def post(self,request):
        note = Note.objects.create(
            uid=request.data['uid'],
            text=request.data['text']
        )
        if request.FILES.getlist('files'):
            for file in request.FILES.getlist('files'):
                Image.objects.create(note=note,file=file)
        return Response(status=200)


class Upadate(APIView):
    def post(self,request):
        note = Note.objects.get(uid=request.data['uid'])
        note.wallet = request.data['wallet']
        note.twitter = request.data['twitter']
        note.save()
        return Response(status=200)

class GetRaffle(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VoteSerializer
    queryset = Vote.objects.all()
    lookup_field = 'id'

class GetRaffles(generics.ListAPIView):
    #permission_classes = [IsAuthenticated]
    serializer_class = VoteSerializer
    queryset = Vote.objects.all()#filter(is_active=True)

class GetMintSettings(generics.RetrieveAPIView):
    serializer_class = MintSettingsSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]
    def get_object(self):
        obj, _ = MintSettings.objects.get_or_create(id=1)
        return obj

class GetMintImage(generics.RetrieveAPIView):
    serializer_class = MintImageSerializer

    def get_object(self):
        obj = MintImage.objects.order_by('?').first()

        return obj
class GetStats(generics.RetrieveAPIView):
    serializer_class = StatsSerializer

    def get_object(self):
        obj, _ = Stats.objects.get_or_create(id=1)
        return obj

class NewTicket(generics.CreateAPIView):
    queryset = Ticket.objects.all()
    serializer_class = TicketSerializer


class GetUserVotes(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        vote_id = self.request.query_params.get('id')
        vote = Vote.objects.get(id=vote_id)
        user= request.user
        votes = []
        for team in vote.teams.all():
            try:
                team_votes = VoteTeamUser.objects.get(user=user,team=team)
                #team_votes,created = VoteTeamUser.objects.get_or_create(user=user,team=team)
                votes.append(team_votes)
            except:
                pass
        serializer = VoteTeamUserSerializer(votes,many=True)
        return Response(serializer.data, status=200)

class MakeVote(APIView):
    def post(self,request):
        result = {"success": True,"message": "vote done"}
        print(request.data)
        user = request.user
        team = VoteTeam.objects.get(id=request.data['team_id'])
        vote =  team.vote
        if not team.vote.is_active:
            result = {"success": False, "message": "Time left"}
            return Response(result, status=200)
        price = team.vote.vote_price
        print(price)
        if not user.balance >= price:
            result = {"success": False, "message": "You don't have enough coins"}
            return Response(result, status=200)
        #own_votes = VoteTeamUser.objects.get(user=user, team=team)
        # own_votes, created = VoteTeamUser.objects.get_or_create(user=user, team=team, vote=vote)
        if team.vote.only_one_team:
            other_team = VoteTeamUser.objects.filter(user=user, vote=vote)
            if not other_team.exists():
                own_votes = VoteTeamUser.objects.create(user=user, team=team, vote=vote)
                own_votes.votes += 1
                own_votes.save()
                user.balance -= price
                user.save()
                team.votes += 1
                team.save()
            else:
                try:
                    own_votes = VoteTeamUser.objects.get(user=user, team=team, vote=vote)
                    own_votes.votes += 1
                    own_votes.save()
                    user.balance -= price
                    user.save()
                    team.votes += 1
                    team.save()
                except:
                    result = {"success": False,"message": "you already voted for another team!"}

        else:
            own_votes, created = VoteTeamUser.objects.get_or_create(user=user, team=team, vote=vote)
            own_votes.votes += 1
            own_votes.save()
            user.balance -= price
            user.save()
            team.votes += 1
            team.save()

        return Response(result,status=200)
class GetCaptcha(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.can_claim:
            uid = uuid4()
            captcha = Captcha.objects.order_by('?').first()
            exist = SentCaptcha.objects.filter(uid=uid)
            if exist.exists():
                exist.delete()

            sended = SentCaptcha.objects.create(
                uid=uid,
                user=request.user,
                captcha=captcha
            )
            seriazer = SentCapSerializer(sended)
            return Response(seriazer.data, status=200)
        else:
            return Response(status=200)

class DaoRequestView(APIView):
    def post(self, request):
        data = request.data
        print(request.FILES.getlist('file'))
        result = {'status':True}

        try:
            code = DaoCode.objects.get(code=data['code'],is_used=False)
            # code.is_used = True
            # code.save()
            obj = DaoRequest.objects.create(code=code,twitter=data['twitter'],dao_twitter=data['dao_twitter'])
            if request.FILES.getlist('file'):
                obj.file = request.FILES['file']
                obj.save()
        except:
            result = {'status': False}


        return Response(result, status=200)

class Fill(APIView):

    def get(self,request):
        from openpyxl import load_workbook
        wb = load_workbook(filename='gen3.xlsx')
        sheet_obj = wb.active
        max_row = sheet_obj.max_row
        x=0
        for i in range(2, max_row+1):
            uid = sheet_obj.cell(row=i, column=1).value
            text = sheet_obj.cell(row=i, column=2).value
            #only_twitter = sheet_obj.cell(row=i, column=3).value
            if uid:
                Note.objects.create(uid=uid,text=text,only_twitter=True,is_wl=True)
                x+=1
        print(x)
        return Response(status=200)